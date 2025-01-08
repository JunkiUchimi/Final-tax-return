import os
from openpyxl import load_workbook
from fpdf import FPDF

class PDF(FPDF):
    def __init__(self):
        super().__init__()
        # 日本語フォントを追加
        self.add_font("NotoSans", fname="NotoSansJP-Regular.ttf", uni=True)
        self.set_font("NotoSans", size=10)

def convert_excel_to_pdf(excel_file, pdf_file):
    """
    指定したExcelファイルをPDFファイルとして保存します。

    Args:
        excel_file (str): Excelファイルのパス
        pdf_file (str): 出力するPDFファイルのパス
    """
    # Excelファイルを読み込む
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook.active  # 最初のシートを選択

    # PDFを生成
    pdf = PDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # シートの内容をPDFに書き込む
    for row in sheet.iter_rows(values_only=True):
        row_data = [str(cell) if cell is not None else "" for cell in row]
        row_text = "  |  ".join(row_data)  # セルを区切り文字で結合
        pdf.cell(0, 10, txt=row_text, ln=True)  # 1行ずつPDFに追加

    # PDFファイルを保存
    pdf.output(pdf_file)
    print(f"PDFファイルとして保存されました: {pdf_file}")

if __name__ == "__main__":
    # Excelファイルの名前と出力するPDFの名前を定義
    excel_file_name = "入力データ.xlsx"
    pdf_file_name = "入力データ.pdf"

    # 現在のスクリプトのディレクトリを取得
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(current_dir, excel_file_name)
    pdf_path = os.path.join(current_dir, pdf_file_name)

    # フォントファイルの存在確認
    font_path = os.path.join(current_dir, "NotoSansJP-Regular.ttf")
    if not os.path.exists(font_path):
        print(f"フォントファイルが見つかりません: {font_path}")
        print("フォントをダウンロードして、スクリプトと同じディレクトリに配置してください。")
        exit()

    # ファイルが存在するか確認して処理を実行
    if os.path.exists(excel_path):
        convert_excel_to_pdf(excel_path, pdf_path)
    else:
        print(f"指定されたExcelファイルが見つかりません: {excel_path}")
